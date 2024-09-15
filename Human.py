import random
import openpyxl

#Human Class
class Human:
    def __init__(self):
        self.name = ""
        self.gender = ""
        self.career = ""
        self.age = 0
        self.priority = 0
        self.generation = ""
    
    def __str__(self):
        return f"\n-Human- \nName: {self.name}\nGender: {self.gender}\nAge: {self.age}\nCareer: {self.career}\nPriority: {self.priority}\nGeneration: {self.generation}"
    
    #Get Name Function
    def getName(self):
        try:
            myNames = []
            workbook = openpyxl.load_workbook('resources/humans.xlsx') #Load File
            sheet = workbook['Sheet1']

            #Get All Names
            for col in sheet.columns:
                if str(col[0].value) == "Name":
                    for name in col[1:]:
                        if name.value != None: #Except None
                            myNames.append(name.value) #Add Name to List
            self.name = random.choice(myNames) #Get Random Name
            return self.name
        except Exception as e:
            print(e)

    #Get Gender Function
    def getGender(self):
        try:
            myGenders = ["Male", "Female", "Other"]
            gender = random.choice(myGenders) #Get Random Gender
            if gender == "Female":
                self.priority += 0.5 #Set Priority For Female
            self.gender = gender
            return self.gender
        except Exception as e:
            print(e)

    #Get Age Function
    def getAge(self):
        try:
            myGeneration = ""
            myAge = 0
            myAge = random.randint(1, 100) #Get Random Age
            #Set Generation
            if myAge > 75:
                myGeneration = "Elder"
                self.generation = myGeneration
                self.age = myAge
                return self.age
            elif myAge > 18:
                myGeneration = "Adult"
                self.generation = myGeneration
                self.age = myAge
                return self.age
            else:
                myGeneration = "Young"
                self.generation = myGeneration
                self.age = myAge
                return self.age
        except Exception as e:
            print(e)

    #Get Career Function
    def getCareer(self):
        try:
            myCareer = ""
            myGeneration = self.generation
            #Set Career
            if myGeneration == "Young":
                myGender = self.gender
                if myGender == "Male":
                    myCareer = "Boy"
                    self.career = myCareer
                    return self.career
                elif myGender == "Female":
                    myCareer = "Girl"
                    self.career = myCareer
                    return self.career
                else:
                    myCareer = "Kid"
                    self.career = myCareer
                    return self.career
            elif myGeneration == "Elder":
                myCareer = "Ederly"
                self.career = myCareer
                return self.career
            else:
                myCareers = []
                workbook = openpyxl.load_workbook('resources/humans.xlsx') #Load File
                sheet = workbook['Sheet1']
                for col in sheet.columns:
                    if str(col[0].value) == "Career": #If Column Name is "Career"
                        for career in col[1:]: #Loop Through All Careers
                            if career.value != None: #Except None
                                myCareers.append(career.value) #Add All Career to List
                myCareer = random.choice(myCareers) #Get Random Career
                workbook.close()
                #Male and Other gender cannot get pregnant !!!
                if myCareer == "Pregnant":
                    if self.gender == "Female":
                        self.career = myCareer
                        return self.career
                    else:
                        careerIndex = myCareers.index(myCareer) #Get Index of "Pregnant"
                        myCareers.pop(careerIndex) #Remove "Pregnant"
                        myCareer = random.choice(myCareers) #reGet Random Career
                        self.career = myCareer
                        return self.career
                else:
                    self.career = myCareer
                    return self.career
    
        except Exception as e:
            print(f"Error: {e}")

    def getPriority(self):
        try:
            if self.generation == "Elder":
                self.priority += 2
                return self.priority
            elif self.generation == "Young":
                self.priority += 3
            else:
                workbook = openpyxl.load_workbook('resources/humans.xlsx')
                sheet = workbook['Sheet1']
                for col in sheet.columns:
                    if str(col[0].value) == "Career":
                        for career in col[1:]:
                            if career.value == self.career:
                                priority = career.offset(0, 1).value
                                self.priority += priority
                return self.priority
        except Exception as e:
            print(f"Error: {e}")

    #Initalize Human
    def getHuman(self):
        self.getName()
        self.getGender()
        self.getAge()
        self.getCareer()
        self.getPriority()
        return self
    

if __name__ == "__main__":
    myhuman = Human()
    print(myhuman.getHuman())
    
